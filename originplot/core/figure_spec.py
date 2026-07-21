from __future__ import annotations

import csv
import json
import math
from pathlib import Path
from typing import Any

from .errors import OriginPlotError


def _number(value: Any, *, column: str, row: int) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise OriginPlotError("E301_FIGURE_DATA_INVALID", f"row {row}: {column} is not numeric") from exc
    if not math.isfinite(number):
        raise OriginPlotError("E301_FIGURE_DATA_INVALID", f"row {row}: {column} is not finite")
    return number


def _read_rows(source: Path, sheet: str | None) -> list[dict[str, Any]]:
    suffix = source.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            dialect = csv.excel_tab if suffix == ".tsv" else csv.excel
            return list(csv.DictReader(handle, dialect=dialect))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise OriginPlotError("E302_XLSX_READER_UNAVAILABLE", "openpyxl is required for XLSX FigureSpecs") from exc
        book = load_workbook(source, read_only=True, data_only=True)
        worksheet = book[sheet] if sheet else book.active
        values = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(values, ())]
        return [dict(zip(headers, row)) for row in values]
    raise OriginPlotError("E303_DATA_FORMAT_UNSUPPORTED", f"only CSV, TSV, and XLSX are supported: {source}")


def load_figure_spec(path: Path) -> dict[str, Any]:
    path = path.resolve()
    try:
        spec = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", f"cannot read FigureSpec: {path}") from exc
    if not isinstance(spec, dict) or spec.get("schema") != "originplot.figurespec.v5":
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", "FigureSpec schema must be originplot.figurespec.v5")
    data_records = spec.get("data")
    plots = spec.get("plots")
    layers = spec.get("layers")
    if (not isinstance(data_records, list) or len(data_records) != 1 or not isinstance(plots, list)
            or len(plots) != 1 or not isinstance(layers, list) or len(layers) != 1):
        raise OriginPlotError("E304_GENERIC_LINE_CONTRACT", "generic_line requires exactly one data record, layer, and plot")
    data = data_records[0]
    plot = plots[0]
    if not isinstance(data, dict) or not isinstance(plot, dict) or str(plot.get("type", "")).lower() != "line":
        raise OriginPlotError("E304_GENERIC_LINE_CONTRACT", "generic_line supports one line plot")
    roles = data.get("roles") or {}
    mapping = plot.get("map") or roles
    x_name, y_name = str(mapping.get("x", "")), str(mapping.get("y", ""))
    if not x_name or not y_name:
        raise OriginPlotError("E305_DATA_ROLES_MISSING", "FigureSpec must name x and y columns")
    source = Path(str(data.get("source", "")))
    if not source.is_absolute():
        source = path.parent / source
    if not source.is_file():
        raise OriginPlotError("E306_DATA_SOURCE_MISSING", f"data source does not exist: {source}")
    rows = _read_rows(source, data.get("sheet") or data.get("worksheet"))
    if not rows or x_name not in rows[0] or y_name not in rows[0]:
        raise OriginPlotError("E307_DATA_COLUMNS_MISSING", f"required columns not found: {x_name}, {y_name}")
    x_values, y_values = [], []
    for index, row in enumerate(rows, start=2):
        x_values.append(_number(row.get(x_name), column=x_name, row=index))
        y_values.append(_number(row.get(y_name), column=y_name, row=index))
    if len(x_values) < 2:
        raise OriginPlotError("E308_DATA_TOO_SHORT", "at least two data rows are required")
    figure = spec.get("figure") or {}
    layer = layers[0]
    return {
        "figure_id": str(figure.get("id") or "figure"),
        "title": str(figure.get("title") or figure.get("id") or "OriginPlot line"),
        "x_label": str((layer.get("x") or {}).get("label") or x_name),
        "y_label": str((layer.get("y") or {}).get("label") or y_name),
        "x_name": x_name,
        "y_name": y_name,
        "x": x_values,
        "y": y_values,
        "source": {"path": str(source), "format": source.suffix.lower().lstrip("."), "data_id": data.get("id")},
        "plot_type": "line",
        "style": plot.get("style") if isinstance(plot.get("style"), dict) else {},
    }
