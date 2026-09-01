from __future__ import annotations

import csv
import hashlib
import json
import math
from pathlib import Path
from typing import Any

from originplot.core.errors import OriginPlotError
from originplot.core.profiles import PROFILE_NAMES

from .models import FIGURE_SPEC_SCHEMA, FigureSpec
from .style import resolve_style

SUPPORTED_TABLE_SUFFIXES = {".csv", ".tsv", ".txt", ".xls", ".xlsx"}
_AXIS_FIELDS = {"title", "unit"}
_PAGE_FIELDS = {"width_cm", "height_cm"}
_VERIFICATION_FIELDS = {
    "profile",
    "require_reopen",
    "require_binding_readback",
    "require_origin_export",
}
_HARD_VERIFICATION_GATES = (
    "require_reopen",
    "require_binding_readback",
    "require_origin_export",
)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_headers(values: tuple[Any, ...] | list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None and str(value).strip() else f"Unnamed_{index}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base}_{count + 1}")
    return headers


def _read_delimited(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".tsv":
        delimiter = "\t"
    elif path.suffix.lower() == ".txt":
        sample = path.read_text(encoding="utf-8-sig", errors="replace")[:8192]
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
        except csv.Error:
            delimiter = "\t" if "\t" in sample else ","
    else:
        delimiter = ","
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        headers = _normalize_headers(next(reader, []))
        return [dict(zip(headers, row)) for row in reader if any(str(value).strip() for value in row)]


def _read_xlsx(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise OriginPlotError("E302_XLSX_READER_UNAVAILABLE", "openpyxl is required for XLSX input") from exc
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        try:
            worksheet = book[sheet] if sheet else book.active
        except KeyError as exc:
            raise OriginPlotError("E309_XLSX_SHEET_MISSING", f"worksheet does not exist: {sheet}") from exc
        values = worksheet.iter_rows(values_only=True)
        headers = _normalize_headers(list(next(values, ())))
        return [dict(zip(headers, row)) for row in values if any(value not in (None, "") for value in row)]
    finally:
        book.close()


def _read_xls(path: Path, sheet: str | None) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise OriginPlotError("E310_XLS_READER_UNAVAILABLE", "xlrd is required for legacy XLS input") from exc
    book = xlrd.open_workbook(str(path), on_demand=True)
    try:
        try:
            worksheet = book.sheet_by_name(sheet) if sheet else book.sheet_by_index(0)
        except (IndexError, xlrd.biffh.XLRDError) as exc:
            raise OriginPlotError("E309_XLSX_SHEET_MISSING", f"worksheet does not exist: {sheet}") from exc
        if worksheet.nrows == 0:
            return []
        headers = _normalize_headers(worksheet.row_values(0))
        return [dict(zip(headers, worksheet.row_values(row))) for row in range(1, worksheet.nrows)]
    finally:
        book.release_resources()


def read_table(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    path = path.resolve()
    if not path.is_file():
        raise OriginPlotError("E306_DATA_SOURCE_MISSING", f"data source does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_TABLE_SUFFIXES:
        raise OriginPlotError("E303_DATA_FORMAT_UNSUPPORTED", f"unsupported table format: {suffix}")
    if suffix in {".csv", ".tsv", ".txt"}:
        return _read_delimited(path)
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet)
    return _read_xls(path, sheet)


def _ensure_object(value: Any, name: str) -> dict[str, Any]:
    if value is None:
        return {}
    if not isinstance(value, dict):
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", f"{name} must be an object")
    return dict(value)


def _normalize_style(value: Any, path: str = "style") -> dict[str, Any]:
    raw_style = _ensure_object(value, path)
    result = resolve_style(user=raw_style)
    rejected = result["rejected"]
    if rejected:
        paths = ", ".join(sorted(f"{path}.{item.get('path') or 'style'}" for item in rejected))
        raise OriginPlotError(
            "E341_STYLE_FIELD_NOT_EXECUTABLE",
            f"style field is not executable by the v6 Origin adapter: {paths}",
        )
    return dict(result["style"])


def _normalize_axis(value: Any, path: str) -> dict[str, Any]:
    axis = _ensure_object(value, path)
    unknown = sorted(set(axis) - _AXIS_FIELDS)
    if unknown:
        raise OriginPlotError(
            "E342_AXIS_CONTRACT_INVALID",
            f"{path}.{unknown[0]} is not executable by the v6 Origin adapter",
        )

    clean: dict[str, Any] = {}
    for field in ("title", "unit"):
        if field not in axis:
            continue
        item = axis[field]
        if not isinstance(item, str):
            raise OriginPlotError("E342_AXIS_CONTRACT_INVALID", f"{path}.{field} must be a string")
        text = item.strip()
        if text:
            clean[field] = text
    if "unit" in clean and "title" not in clean:
        raise OriginPlotError("E342_AXIS_CONTRACT_INVALID", f"{path}.unit requires title")
    return clean


def _normalize_panel(panel: Any, index: int) -> dict[str, Any]:
    path = f"figure.panels.{index}"
    item = _ensure_object(panel, path)
    child_figure = _ensure_object(item.get("figure"), f"{path}.figure")
    child_type = str(child_figure.get("type") or "").strip().lower()
    if not child_type:
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", f"{path}.figure.type is required")
    if child_type == "multi_panel":
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", f"{path} cannot contain nested multi_panel")

    if "x_axis" in child_figure:
        child_figure["x_axis"] = _normalize_axis(child_figure.get("x_axis"), f"{path}.figure.x_axis")
    if "y_axis" in child_figure:
        child_figure["y_axis"] = _normalize_axis(child_figure.get("y_axis"), f"{path}.figure.y_axis")

    clean = dict(item)
    clean["figure"] = child_figure
    clean["data"] = _ensure_object(item.get("data"), f"{path}.data")
    if "style" in item:
        clean["style"] = _normalize_style(item.get("style"), f"{path}.style")
    if item.get("layout"):
        raise OriginPlotError(
            "E343_LAYOUT_CONTRACT_INVALID",
            f"{path}: panel-specific layout is not compiled by v6 multi_panel",
        )
    clean.pop("layout", None)
    return clean


def _normalize_figure(value: Any) -> dict[str, Any]:
    figure = _ensure_object(value, "figure")
    plot_type = str(figure.get("type") or "").strip().lower()
    if not plot_type:
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", "figure.type is required")
    figure["type"] = plot_type

    if plot_type == "multi_panel":
        if figure.get("x_axis") or figure.get("y_axis"):
            raise OriginPlotError(
                "E342_AXIS_CONTRACT_INVALID",
                "figure.x_axis/y_axis are not executable on the multi_panel container; put axes on each panel",
            )
        panels = figure.get("panels")
        if not isinstance(panels, list):
            raise OriginPlotError("E300_FIGURE_SPEC_INVALID", "figure.panels must be a list")
        figure["panels"] = [_normalize_panel(panel, index) for index, panel in enumerate(panels)]
        return figure

    if "x_axis" in figure:
        figure["x_axis"] = _normalize_axis(figure.get("x_axis"), "figure.x_axis")
    if "y_axis" in figure:
        figure["y_axis"] = _normalize_axis(figure.get("y_axis"), "figure.y_axis")
    return figure


def _positive_number(value: Any, path: str) -> float:
    if not isinstance(value, (int, float)) or isinstance(value, bool) or not math.isfinite(float(value)) or float(value) <= 0:
        raise OriginPlotError("E343_LAYOUT_CONTRACT_INVALID", f"{path} must be a positive number")
    return float(value)


def _normalize_layout(value: Any, plot_type: str) -> dict[str, Any]:
    layout = _ensure_object(value, "layout")
    unknown = sorted(set(layout) - {"page", "panels"})
    if unknown:
        raise OriginPlotError(
            "E343_LAYOUT_CONTRACT_INVALID",
            f"layout.{unknown[0]} is not executable by the v6 compiler",
        )

    clean: dict[str, Any] = {}
    if "page" in layout:
        page = _ensure_object(layout.get("page"), "layout.page")
        page_unknown = sorted(set(page) - _PAGE_FIELDS)
        if page_unknown:
            raise OriginPlotError(
                "E343_LAYOUT_CONTRACT_INVALID",
                f"layout.page.{page_unknown[0]} is not executable by the v6 Origin adapter",
            )
        has_width = "width_cm" in page and page.get("width_cm") is not None
        has_height = "height_cm" in page and page.get("height_cm") is not None
        if has_width != has_height:
            raise OriginPlotError(
                "E343_LAYOUT_CONTRACT_INVALID",
                "layout.page.width_cm and height_cm must be provided together",
            )
        if has_width and has_height:
            clean["page"] = {
                "width_cm": _positive_number(page["width_cm"], "layout.page.width_cm"),
                "height_cm": _positive_number(page["height_cm"], "layout.page.height_cm"),
            }
        elif page:
            clean["page"] = {}

    if "panels" in layout:
        if plot_type != "multi_panel":
            raise OriginPlotError(
                "E343_LAYOUT_CONTRACT_INVALID",
                "layout.panels is only valid for the compile-only multi_panel primitive",
            )
        clean["panels"] = _ensure_object(layout.get("panels"), "layout.panels")
    return clean


def _normalize_verification(value: Any) -> dict[str, Any]:
    verification = _ensure_object(value, "verification")
    unknown = sorted(set(verification) - _VERIFICATION_FIELDS)
    if unknown:
        raise OriginPlotError(
            "E344_VERIFICATION_CONTRACT_INVALID",
            f"verification.{unknown[0]} is not a supported v6 verification field",
        )

    profile = str(verification.get("profile") or "standard").strip().lower()
    if profile not in PROFILE_NAMES:
        raise OriginPlotError(
            "E344_VERIFICATION_CONTRACT_INVALID",
            f"verification.profile must be one of {', '.join(PROFILE_NAMES)}",
        )
    for field in _HARD_VERIFICATION_GATES:
        if field in verification and verification[field] is not True:
            raise OriginPlotError(
                "E344_VERIFICATION_CONTRACT_INVALID",
                f"verification.{field} is a mandatory v6 gate and cannot be disabled",
            )
    return {
        "profile": profile,
        "require_reopen": True,
        "require_binding_readback": True,
        "require_origin_export": True,
    }


def _mapped_columns(data: dict[str, Any], figure: dict[str, Any]) -> set[str]:
    columns: set[str] = set()
    plot_type = str(figure.get("type") or "").lower()
    if plot_type == "multi_panel":
        for panel in figure.get("panels") or []:
            if isinstance(panel, dict):
                columns.update(
                    _mapped_columns(
                        _ensure_object(panel.get("data"), "panel.data"),
                        _ensure_object(panel.get("figure"), "panel.figure"),
                    )
                )
        return columns
    for item in data.get("series") or []:
        if isinstance(item, dict):
            for key in ("x", "y", "x_error", "y_error", "category", "group", "label", "z"):
                value = item.get(key)
                if isinstance(value, str) and value.strip():
                    columns.add(value.strip())
    matrix = data.get("matrix")
    if isinstance(matrix, dict):
        for key in ("x", "y", "z"):
            value = matrix.get(key)
            if isinstance(value, str) and value.strip():
                columns.add(value.strip())
    return columns


def normalize_figure_spec(payload: dict[str, Any], base_dir: Path | None = None) -> FigureSpec:
    if not isinstance(payload, dict) or payload.get("schema") != FIGURE_SPEC_SCHEMA:
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", f"FigureSpec schema must be {FIGURE_SPEC_SCHEMA}")
    base_dir = (base_dir or Path.cwd()).resolve()
    source = _ensure_object(payload.get("source"), "source")
    raw_file = source.get("file")
    if not isinstance(raw_file, str) or not raw_file.strip():
        raise OriginPlotError("E306_DATA_SOURCE_MISSING", "FigureSpec source.file is required")
    source_path = Path(raw_file)
    if not source_path.is_absolute():
        source_path = base_dir / source_path
    source_path = source_path.resolve()
    if not source_path.is_file():
        raise OriginPlotError("E306_DATA_SOURCE_MISSING", f"data source does not exist: {source_path}")
    if source_path.suffix.lower() not in SUPPORTED_TABLE_SUFFIXES:
        raise OriginPlotError("E303_DATA_FORMAT_UNSUPPORTED", f"unsupported table format: {source_path.suffix}")

    actual_hash = file_sha256(source_path)
    expected_hash = str(source.get("hash") or "").strip().lower()
    if expected_hash and expected_hash != actual_hash:
        raise OriginPlotError("E311_SOURCE_HASH_MISMATCH", "FigureSpec source hash does not match the current data file")

    data = _ensure_object(payload.get("data"), "data")
    figure = _normalize_figure(payload.get("figure"))
    style = _normalize_style(payload.get("style"), "style")
    plot_type = str(figure["type"])
    layout = _normalize_layout(payload.get("layout"), plot_type)
    verification = _normalize_verification(payload.get("verification"))

    sheet = str(source.get("sheet") or "").strip() or None
    rows = read_table(source_path, sheet)
    if not rows:
        raise OriginPlotError("E308_DATA_TOO_SHORT", "source table contains no data rows")
    headers = set(rows[0])
    missing = sorted(_mapped_columns(data, figure) - headers)
    if missing:
        raise OriginPlotError("E307_DATA_COLUMNS_MISSING", "required columns not found: " + ", ".join(missing))

    raw = dict(payload)
    raw["figure"] = figure
    raw["layout"] = layout
    raw["verification"] = verification
    if "style" in payload or style:
        raw["style"] = style

    return FigureSpec(
        source_path=source_path,
        source_hash=actual_hash,
        sheet=sheet,
        data=data,
        figure=figure,
        style=style,
        layout=layout,
        verification=verification,
        raw=raw,
    )


def load_figure_spec(path: Path) -> FigureSpec:
    path = path.resolve()
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", f"cannot read FigureSpec: {path}") from exc
    if not isinstance(payload, dict):
        raise OriginPlotError("E300_FIGURE_SPEC_INVALID", "FigureSpec root must be an object")
    return normalize_figure_spec(payload, path.parent)
