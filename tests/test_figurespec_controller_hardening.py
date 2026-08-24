from __future__ import annotations

from pathlib import Path

import pytest

import originplot.controller as controller
from originplot.core.errors import OriginPlotError
from originplot.core.figure_spec import _read_rows


class _FakeWorksheet:
    def iter_rows(self, values_only: bool = True):
        assert values_only is True
        yield ("x", "y")
        yield (1, 2)


class _FakeWorkbook:
    def __init__(self) -> None:
        self.active = _FakeWorksheet()
        self.closed = False

    def __getitem__(self, name: str):
        if name != "Data":
            raise KeyError(name)
        return self.active

    def close(self) -> None:
        self.closed = True


def test_xlsx_reader_closes_workbook(monkeypatch: pytest.MonkeyPatch) -> None:
    import openpyxl

    book = _FakeWorkbook()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: book)

    rows = _read_rows(Path("sample.xlsx"), "Data")

    assert rows == [{"x": 1, "y": 2}]
    assert book.closed is True


def test_xlsx_missing_sheet_is_stable_error_and_closes(monkeypatch: pytest.MonkeyPatch) -> None:
    import openpyxl

    book = _FakeWorkbook()
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *args, **kwargs: book)

    with pytest.raises(OriginPlotError) as exc_info:
        _read_rows(Path("sample.xlsx"), "Missing")

    assert exc_info.value.code == "E309_XLSX_SHEET_MISSING"
    assert book.closed is True


def test_powershell_resolver_falls_back_to_windows_powershell(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        controller.shutil,
        "which",
        lambda name: r"C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe" if name == "powershell" else None,
    )

    executable = controller._resolve_powershell_executable()

    assert executable.name.lower() == "powershell.exe"


def test_powershell_resolver_reports_environment_mismatch(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(controller.shutil, "which", lambda name: None)
    monkeypatch.setattr(controller.Path, "is_file", lambda self: False)

    with pytest.raises(OriginPlotError) as exc_info:
        controller._resolve_powershell_executable()

    assert exc_info.value.code == "E120_ENVIRONMENT_MISMATCH"
